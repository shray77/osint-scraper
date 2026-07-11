"""
reporters.py — генерация отчётов из собранных данных.

MarkdownReporter — человекочитаемый .md отчёт в стиле презентации-примера.
ExcelReporter    — многолистовой .xlsx: Профиль / Метрики / Рынок / Источники / Логи.
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("osint.reporters")


# ---------------------------------------------------------------------------
# Markdown reporter
# ---------------------------------------------------------------------------
class MarkdownReporter:
    """Генерирует Markdown-отчёт по структуре презентации-примера."""

    def render(self, report: dict) -> str:
        lines = []
        company = report.get("company", {})
        market = report.get("market", {})

        # Заголовок
        title = company.get("full_name") or company.get("short_name") or report.get("company_input", "—")
        lines.append(f"# {title}")
        lines.append(f"**OSINT-отчёт · {datetime.now().strftime('%Y-%m-%d %H:%M')}**")
        lines.append("")
        lines.append(f"*Запрос:* `{report.get('company_input', '')}`")
        if company.get("inn"):
            lines.append(f"*ИНН:* `{company['inn']}`  ·  *ОГРН:* `{company.get('ogrn', '—')}`")
        lines.append("")
        lines.append("---")
        lines.append("")

        # 1. О компании
        lines.append("## 1. О компании")
        lines.append("")
        facts = []
        if company.get("full_name"):
            facts.append(("Полное наименование", company["full_name"]))
        if company.get("short_name"):
            facts.append(("Сокращённое наименование", company["short_name"]))
        if company.get("registration_date"):
            facts.append(("Дата регистрации", company["registration_date"]))
        if company.get("legal_address"):
            facts.append(("Юридический адрес", company["legal_address"]))
        if company.get("status"):
            facts.append(("Статус", company["status"]))
        if company.get("director"):
            facts.append(("Руководитель", company["director"]))
        if company.get("founder"):
            facts.append(("Учредитель", company["founder"]))
        if company.get("okved_main"):
            facts.append(("Основной ОКВЭД", company["okved_main"]))
            industry = company.get("industry") or ""
            if industry:
                facts.append(("Отрасль (по ОКВЭД)", industry))
        if company.get("authorized_capital"):
            facts.append(("Уставный капитал", company["authorized_capital"]))

        if facts:
            for k, v in facts:
                lines.append(f"- **{k}:** {v}")
        else:
            lines.append("_Данные об идентификации компании не найдены._")
        lines.append("")

        # Все ОКВЭД
        okved_all = company.get("okved_all") or []
        if okved_all:
            lines.append("### Все ОКВЭД")
            lines.append("")
            for o in okved_all[:20]:
                lines.append(f"- {o}")
            lines.append("")

        # Все совпадения юрлиц (из EGRUL) — для disambiguation
        all_matches = company.get("all_matches") or []
        if all_matches and len(all_matches) > 1:
            lines.append("### Все совпадения по запросу (для disambiguation)")
            lines.append("")
            lines.append("> ⚠️ **Важно:** у холдинга может быть несколько юрлиц с похожим названием. "
                         "Проверьте, что выбрали правильное — ОКВЭД у них может отличаться "
                         "(например, управляющая компания vs. производственное юрлицо).")
            lines.append("")
            for i, m in enumerate(all_matches, 1):
                status_emoji = "✅" if m.get("status") == "действует" else "❌"
                lines.append(f"**{i}. {m.get('short_name') or m.get('full_name', '—')}** {status_emoji}")
                lines.append(f"   - ИНН: `{m.get('inn', '—')}` · ОГРН: `{m.get('ogrn', '—')}` · КПП: `{m.get('kpp', '—')}`")
                lines.append(f"   - Регион: {m.get('region', '—')}")
                lines.append(f"   - Регистрация: {m.get('registration_date', '—')}")
                if m.get('director_role') and m.get('director'):
                    lines.append(f"   - {m.get('director_role')}: {m.get('director')}")
                if m.get('liquidation_date'):
                    lines.append(f"   - Ликвидация: {m.get('liquidation_date')}")
                lines.append("")

        # 2. Ключевые метрики (финансы — базово + пометка)
        lines.append("## 2. Ключевые метрики (базово)")
        lines.append("")
        metrics = []
        if company.get("revenue"):
            metrics.append(("Выручка", company["revenue"]))
        if company.get("profit"):
            metrics.append(("Чистая прибыль", company["profit"]))
        if company.get("assets"):
            metrics.append(("Стоимость активов", company["assets"]))
        if company.get("employees"):
            metrics.append(("Численность сотрудников", company["employees"]))

        if metrics:
            for k, v in metrics:
                lines.append(f"- **{k}:** {v}")
        else:
            lines.append("_Финансовые показатели не найдены в открытых реестрах._")
        lines.append("")

        caveat = company.get("finance_caveat") or report.get("finance_caveat")
        if caveat:
            lines.append(f"> ⚠️ **Пометка по финансам:** {caveat}")
            lines.append("")

        # 3. Рынок и доля
        lines.append("## 3. Рынок и доля")
        lines.append("")
        if market.get("industry"):
            lines.append(f"**Отрасль для анализа рынка:** {market['industry']}")
            lines.append("")

        # Объём рынка
        sizes = market.get("market_size_candidates") or []
        if sizes:
            lines.append("### Объём рынка (кандидаты из поисковых сниппетов)")
            lines.append("")
            for s in sizes[:5]:
                lines.append(f"- **{s.get('value', '—')}** ({s.get('year', '—')}) — {s.get('text', '')[:150]}  ")
                lines.append(f"  *Источник:* {s.get('source', '')}")
            lines.append("")
        else:
            lines.append("_Данные по объёму рынка не найдены._")
            lines.append("")

        # Динамика
        dyn = market.get("dynamics_candidates") or []
        if dyn:
            lines.append("### Динамика рынка")
            lines.append("")
            for d in dyn[:3]:
                lines.append(f"- {d.get('value', '—')}: {d.get('text', '')[:200]}  ")
                lines.append(f"  *Источник:* {d.get('source', '')}")
            lines.append("")

        # Топ-игроки
        tops = market.get("top_players_candidates") or []
        if tops:
            lines.append("### ТОП-игроки и лидеры")
            lines.append("")
            for t in tops[:3]:
                lines.append(f"- {t.get('text', '')[:250]}")
                lines.append(f"  *Источник:* {t.get('source', '')}")
            lines.append("")

        # Позиция запрошенной компании
        positions = market.get("company_position_candidates") or []
        if positions:
            lines.append("### Позиция компании на рынке")
            lines.append("")
            for p in positions[:3]:
                lines.append(f"- {p.get('text', '')[:250]}")
                lines.append(f"  *Источник:* {p.get('source', '')}")
            lines.append("")

        # Сайт компании (если нашли)
        site = report.get("company_site") or {}
        if site and site.get("site_url"):
            lines.append("## 4. Сайт компании")
            lines.append("")
            lines.append(f"**URL:** [{site.get('site_url', '')}]({site.get('site_url', '')})")
            if site.get("site_title"):
                lines.append(f"**Заголовок страницы:** {site['site_title']}")
            lines.append("")

            # Контакты
            phones = site.get("phones") or []
            emails = site.get("emails") or []
            if phones or emails:
                lines.append("### Контакты с сайта")
                lines.append("")
                if phones:
                    lines.append(f"- **Телефоны:** {', '.join(phones)}")
                if emails:
                    lines.append(f"- **Email:** {', '.join(emails)}")
                lines.append("")

            # Соцсети
            socials = site.get("socials") or {}
            if socials:
                lines.append("### Соцсети")
                lines.append("")
                for name, urls in socials.items():
                    lines.append(f"- **{name}**: {', '.join(urls)}")
                lines.append("")

            # Интересные ссылки
            links = site.get("interesting_links") or []
            if links:
                lines.append("### Ключевые страницы сайта")
                lines.append("")
                for l in links[:10]:
                    lines.append(f"- [{l.get('text', '')}]({l.get('url', '')})")
                lines.append("")

        # 5. Источники
        lines.append("## 5. Источники")
        lines.append("")
        sources = report.get("sources_used") or []
        if sources:
            for s in sources:
                lines.append(f"- {s}")
        else:
            lines.append("_Источники не зафиксированы._")
        lines.append("")

        # 5. Ошибки (прозрачность)
        errors = report.get("errors") or []
        if errors:
            lines.append("## 6. Ошибки и ограничения")
            lines.append("")
            for e in errors:
                lines.append(f"- {e}")
            lines.append("")

        # 7. Что добавить для полной картины
        lines.append("## 7. Рекомендации по расширению")
        lines.append("")
        lines.append("Для более полной картины в следующих версиях скраппера рекомендуется подключить:")
        lines.append("- Скрапинг сайта компании (каталог, цены, адреса магазинов)")
        lines.append("- Yandex Maps / 2GIS (география точек, отзывы, рейтинги)")
        lines.append("- Маркетплейсы Ozon / Wildberries (SKU, рейтинг продавца)")
        lines.append("- Новости и СМИ (упоминания, ключевые темы)")
        lines.append("- Соцсети VK / Telegram (подписчики, активность)")
        lines.append("")
        lines.append("---")
        lines.append(f"*Отчёт сгенерирован OSINT-скраппером · {datetime.now().isoformat(timespec='seconds')}*")

        return "\n".join(lines)


# ---------------------------------------------------------------------------
# Excel reporter
# ---------------------------------------------------------------------------
class ExcelReporter:
    """Многолистовой .xlsx: Профиль / Метрики / Рынок / Источники / Логи."""

    HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    def render(self, report: dict) -> Workbook:
        wb = Workbook()

        self._sheet_profile(wb, report)
        self._sheet_matches(wb, report)
        self._sheet_market(wb, report)
        self._sheet_okved(wb, report)
        self._sheet_sources(wb, report)
        self._sheet_logs(wb, report)

        return wb

    def _style_header(self, ws, row=1):
        for cell in ws[row]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = self.BORDER
        ws.row_dimensions[row].height = 24

    def _autosize(self, ws, max_width=80):
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            length = max(
                (len(str(c.value)) for c in col if c.value is not None),
                default=10,
            )
            ws.column_dimensions[letter].width = min(max_width, max(12, length + 2))

    def _sheet_profile(self, wb: Workbook, report: dict):
        ws = wb.active
        ws.title = "Профиль"
        company = report.get("company", {})

        ws["A1"] = "Поле"
        ws["B1"] = "Значение"
        self._style_header(ws, 1)

        rows = [
            ("Запрос", report.get("company_input", "")),
            ("Полное наименование", company.get("full_name", "")),
            ("Сокращённое наименование", company.get("short_name", "")),
            ("ИНН", company.get("inn", "")),
            ("ОГРН", company.get("ogrn", "")),
            ("КПП", company.get("kpp", "")),
            ("Дата регистрации", company.get("registration_date", "")),
            ("Юридический адрес", company.get("legal_address", "")),
            ("Статус", company.get("status", "")),
            ("Руководитель", company.get("director", "")),
            ("Учредитель", company.get("founder", "")),
            ("Основной ОКВЭД", company.get("okved_main", "")),
            ("Отрасль (по ОКВЭД)", company.get("industry", "")),
            ("Уставный капитал", company.get("authorized_capital", "")),
            ("Выручка", company.get("revenue", "")),
            ("Чистая прибыль", company.get("profit", "")),
            ("Стоимость активов", company.get("assets", "")),
            ("Численность сотрудников", company.get("employees", "")),
            ("Пометка по финансам", company.get("finance_caveat", "")),
            ("Сгенерировано", report.get("generated_at", "")),
        ]
        for i, (k, v) in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=k).border = self.BORDER
            ws.cell(row=i, column=2, value=v or "").border = self.BORDER
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 80

    def _sheet_matches(self, wb: Workbook, report: dict):
        """Лист «Все совпадения юрлиц» — для disambiguation."""
        ws = wb.create_sheet("Все совпадения")
        ws["A1"] = "№"
        ws["B1"] = "Краткое название"
        ws["C1"] = "Полное название"
        ws["D1"] = "ИНН"
        ws["E1"] = "ОГРН"
        ws["F1"] = "КПП"
        ws["G1"] = "Регион"
        ws["H1"] = "Регистрация"
        ws["I1"] = "Должность руководителя"
        ws["J1"] = "Руководитель"
        ws["K1"] = "Статус"
        self._style_header(ws, 1)
        matches = report.get("company", {}).get("all_matches") or []
        for i, m in enumerate(matches, start=1):
            r = i + 1
            ws.cell(row=r, column=1, value=i)
            ws.cell(row=r, column=2, value=m.get("short_name", ""))
            ws.cell(row=r, column=3, value=m.get("full_name", ""))
            ws.cell(row=r, column=4, value=m.get("inn", ""))
            ws.cell(row=r, column=5, value=m.get("ogrn", ""))
            ws.cell(row=r, column=6, value=m.get("kpp", ""))
            ws.cell(row=r, column=7, value=m.get("region", ""))
            ws.cell(row=r, column=8, value=m.get("registration_date", ""))
            ws.cell(row=r, column=9, value=m.get("director_role", ""))
            ws.cell(row=r, column=10, value=m.get("director", ""))
            ws.cell(row=r, column=11, value=m.get("status", ""))
            for c in range(1, 12):
                ws.cell(row=r, column=c).border = self.BORDER
                ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
        widths = [4, 30, 50, 14, 16, 12, 25, 14, 28, 30, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _sheet_market(self, wb: Workbook, report: dict):
        ws = wb.create_sheet("Рынок")
        market = report.get("market", {})

        ws["A1"] = "Отрасль"
        ws["B1"] = "Показатель"
        ws["C1"] = "Значение"
        ws["D1"] = "Год"
        ws["E1"] = "Источник"
        ws["F1"] = "Контекст (сниппет)"
        self._style_header(ws, 1)

        row = 2
        ws.cell(row=row, column=1, value=market.get("industry", ""))
        ws.cell(row=row, column=2, value="—")
        ws.cell(row=row, column=3, value="—")
        ws.cell(row=row, column=4, value="—")
        ws.cell(row=row, column=5, value="—")
        ws.cell(row=row, column=6, value="Обзор рынка")
        row += 1

        for s in market.get("market_size_candidates", []):
            ws.cell(row=row, column=1, value=market.get("industry", ""))
            ws.cell(row=row, column=2, value="Объём рынка")
            ws.cell(row=row, column=3, value=s.get("value", ""))
            ws.cell(row=row, column=4, value=s.get("year", ""))
            ws.cell(row=row, column=5, value=s.get("source", ""))
            ws.cell(row=row, column=6, value=s.get("text", "")[:300])
            row += 1

        for d in market.get("dynamics_candidates", []):
            ws.cell(row=row, column=1, value=market.get("industry", ""))
            ws.cell(row=row, column=2, value="Динамика")
            ws.cell(row=row, column=3, value=d.get("value", ""))
            ws.cell(row=row, column=4, value="—")
            ws.cell(row=row, column=5, value=d.get("source", ""))
            ws.cell(row=row, column=6, value=d.get("text", "")[:300])
            row += 1

        for t in market.get("top_players_candidates", []):
            ws.cell(row=row, column=1, value=market.get("industry", ""))
            ws.cell(row=row, column=2, value="ТОП-игроки")
            ws.cell(row=row, column=3, value="—")
            ws.cell(row=row, column=4, value="—")
            ws.cell(row=row, column=5, value=t.get("source", ""))
            ws.cell(row=row, column=6, value=t.get("text", "")[:300])
            row += 1

        for p in market.get("company_position_candidates", []):
            ws.cell(row=row, column=1, value=market.get("industry", ""))
            ws.cell(row=row, column=2, value="Позиция компании")
            ws.cell(row=row, column=3, value="—")
            ws.cell(row=row, column=4, value="—")
            ws.cell(row=row, column=5, value=p.get("source", ""))
            ws.cell(row=row, column=6, value=p.get("text", "")[:300])
            row += 1

        for r in range(2, row):
            for c in range(1, 7):
                ws.cell(row=r, column=c).border = self.BORDER
                ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 40
        ws.column_dimensions["F"].width = 60

    def _sheet_okved(self, wb: Workbook, report: dict):
        ws = wb.create_sheet("ОКВЭД")
        ws["A1"] = "№"
        ws["B1"] = "Описание ОКВЭД"
        self._style_header(ws, 1)
        okveds = report.get("company", {}).get("okved_all") or []
        for i, o in enumerate(okveds, start=1):
            ws.cell(row=i + 1, column=1, value=i).border = self.BORDER
            ws.cell(row=i + 1, column=2, value=o).border = self.BORDER
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 80

    def _sheet_sources(self, wb: Workbook, report: dict):
        ws = wb.create_sheet("Источники")
        ws["A1"] = "№"
        ws["B1"] = "URL"
        self._style_header(ws, 1)
        sources = report.get("sources_used") or []
        for i, s in enumerate(sources, start=1):
            ws.cell(row=i + 1, column=1, value=i).border = self.BORDER
            ws.cell(row=i + 1, column=2, value=s).border = self.BORDER
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 100

    def _sheet_logs(self, wb: Workbook, report: dict):
        ws = wb.create_sheet("Логи")
        ws["A1"] = "Тип"
        ws["B1"] = "Сообщение"
        self._style_header(ws, 1)
        row = 2
        for e in report.get("errors", []) or []:
            ws.cell(row=row, column=1, value="error").border = self.BORDER
            ws.cell(row=row, column=2, value=e).border = self.BORDER
            row += 1
        if row == 2:
            ws.cell(row=2, column=1, value="info").border = self.BORDER
            ws.cell(row=2, column=2, value="Ошибок нет").border = self.BORDER
            row += 1
        for k, v in (report.get("collector_stats", {}) or {}).items():
            ws.cell(row=row, column=1, value="stats").border = self.BORDER
            ws.cell(row=row, column=2, value=f"{k}: {v}").border = self.BORDER
            row += 1
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 100
